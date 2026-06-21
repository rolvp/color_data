import math
import os
import uuid
from collections import defaultdict
from datetime import datetime, timezone, date
from typing import Optional, List, Tuple

from sqlalchemy.orm import Session
from sqlalchemy import or_, func, and_
from fastapi import HTTPException, UploadFile

from database import beijing_now

from models import Sample, ColorMeasurement, Photo
from schemas import SampleCreate, SampleUpdate, SampleListItem, MeasurementCreate, MeasurementUpdate


def _normalize_device(device: Optional[str]) -> str:
    value = (device or "SP64").strip()
    return value or "SP64"


def _normalize_angle(angle: Optional[str]) -> str:
    return (angle or "").strip()


def _normalize_aging_hours(aging_hours: Optional[float]) -> float:
    try:
        return float(aging_hours or 0)
    except (TypeError, ValueError):
        return 0.0


def _measurement_key(sample_id: int, device: Optional[str], angle: Optional[str], aging_hours: Optional[float]) -> tuple:
    return (
        sample_id,
        _normalize_device(device),
        _normalize_angle(angle),
        _normalize_aging_hours(aging_hours),
    )


def _measurement_sort_key(measurement: ColorMeasurement) -> tuple:
    min_dt = datetime.min
    measurement_date = measurement.measurement_date or min_dt
    created_at = measurement.created_at or min_dt
    return (measurement_date, created_at, measurement.id or 0)


def _has_exact_duplicate_measurement(
    existing_measurements: List[ColorMeasurement],
    sample_id: int,
    device: Optional[str],
    angle: Optional[str],
    aging_hours: Optional[float],
    L: float,
    a: float,
    b: float,
    notes: Optional[str],
) -> bool:
    expected_key = _measurement_key(sample_id, device, angle, aging_hours)
    expected_signature = _measurement_signature(L, a, b, notes)
    for measurement in existing_measurements:
        if _measurement_key(sample_id, measurement.device, measurement.angle, measurement.aging_hours) != expected_key:
            continue
        if _measurement_signature(measurement.L, measurement.a, measurement.b, measurement.notes) == expected_signature:
            return True
    return False


def _measurement_signature(L: float, a: float, b: float, notes: Optional[str]) -> tuple:
    return (
        round(float(L), 6),
        round(float(a), 6),
        round(float(b), 6),
        (notes or "").strip(),
    )


def _select_latest_measurements(measurements: List[ColorMeasurement]) -> List[ColorMeasurement]:
    latest_by_key = {}
    for measurement in sorted(measurements, key=_measurement_sort_key, reverse=True):
        key = _measurement_key(
            measurement.sample_id,
            measurement.device,
            measurement.angle,
            measurement.aging_hours,
        )
        if key not in latest_by_key:
            latest_by_key[key] = measurement

    return sorted(
        latest_by_key.values(),
        key=lambda measurement: (
            _normalize_aging_hours(measurement.aging_hours),
            _measurement_sort_key(measurement),
        ),
    )


def _get_visible_measurements_for_sample(db: Session, sample_id: int) -> List[ColorMeasurement]:
    measurements = (
        db.query(ColorMeasurement)
        .filter(ColorMeasurement.sample_id == sample_id)
        .all()
    )
    return _select_latest_measurements(measurements)


# ============================================================
# Delta-E Calculation (CIE76)
# ============================================================

def calculate_delta_e_cie76(L1: float, a1: float, b1: float,
                             L2: float, a2: float, b2: float) -> float:
    """Calculate CIE76 Delta-E between two LAB colors."""
    delta_L = L1 - L2
    delta_a = a1 - a2
    delta_b = b1 - b2
    return round(math.sqrt(delta_L ** 2 + delta_a ** 2 + delta_b ** 2), 2)


def get_baseline_measurements(db: Session, sample_id: int) -> List[ColorMeasurement]:
    """Get all baseline measurements (aging_hours == 0) for a sample.
    Returns list - for MT12 there will be 6 baselines (one per angle)."""
    return [
        measurement
        for measurement in _get_visible_measurements_for_sample(db, sample_id)
        if _normalize_aging_hours(measurement.aging_hours) == 0
    ]


def get_baseline_for_measurement(db: Session, measurement: ColorMeasurement) -> Optional[ColorMeasurement]:
    """Get the baseline for a specific measurement (same device + angle, aging_hours=0)."""
    for baseline in get_baseline_measurements(db, measurement.sample_id):
        if _normalize_device(baseline.device) != _normalize_device(measurement.device):
            continue
        if _normalize_angle(baseline.angle) != _normalize_angle(measurement.angle):
            continue
        return baseline
    return None


def recalculate_delta_e_for_sample(db: Session, sample_id: int):
    """Recalculate delta_E: aging_hours=0 are baselines (delta_E=null).
    Non-baseline measurements compare against same-device+angle baseline."""
    baselines = get_baseline_measurements(db, sample_id)
    baseline_ids = {b.id for b in baselines}

    # Build a lookup: (device, angle) -> baseline
    baseline_map = {}
    for b in baselines:
        key = (_normalize_device(b.device), _normalize_angle(b.angle))
        if key not in baseline_map:
            baseline_map[key] = b

    measurements = (
        db.query(ColorMeasurement)
        .filter(ColorMeasurement.sample_id == sample_id)
        .order_by(ColorMeasurement.aging_hours, ColorMeasurement.id)
        .all()
    )
    for m in measurements:
        if m.id in baseline_ids:
            m.delta_E = None
        else:
            key = (_normalize_device(m.device), _normalize_angle(m.angle))
            baseline = baseline_map.get(key)
            if not baseline:
                # Fallback: use any baseline from same device
                for bk, bv in baseline_map.items():
                    if bk[0] == _normalize_device(m.device):
                        baseline = bv
                        break
            if not baseline and baseline_map:
                baseline = list(baseline_map.values())[0]
            if baseline:
                m.delta_E = calculate_delta_e_cie76(
                    baseline.L, baseline.a, baseline.b, m.L, m.a, m.b
                )
            else:
                m.delta_E = None
    db.commit()


# ============================================================
# Sample CRUD
# ============================================================

def _sample_to_list_item(sample: Sample) -> SampleListItem:
    """Convert a Sample ORM object to SampleListItem with computed fields."""
    try:
        measurements = _select_latest_measurements(sample.measurements or [])
        measurement_count = len(measurements)
    except Exception:
        measurements = []
        measurement_count = 0

    latest_delta_e = None
    if measurement_count > 0:
        try:
            prioritized_measurements = [
                measurement
                for measurement in measurements
                if _normalize_angle(measurement.angle) == "r45as45"
            ]
            target_measurements = prioritized_measurements or measurements
            latest = max(target_measurements, key=_measurement_sort_key)
            latest_delta_e = latest.delta_E
        except Exception:
            latest_delta_e = None

    try:
        p_count = len(sample.photos) if hasattr(sample, 'photos') and sample.photos is not None else 0
    except Exception:
        p_count = 0

    return SampleListItem(
        id=sample.id,
        name=sample.name,
        code=sample.code,
        category=sample.category or "",
        brand=sample.brand or "",
        model=sample.model or "",
        color_name=sample.color_name or "",
        other_info=sample.other_info or "",
        test_condition=sample.test_condition or "",
        aging_time=sample.aging_time or "",
        device_info=sample.device_info or "",
        test_device=sample.test_device or "",
        measurement_test=sample.measurement_test or "",
        description=sample.description or "",
        measurement_count=measurement_count,
        latest_delta_e=latest_delta_e,
        photo_count=p_count,
        created_at=sample.created_at or datetime.now(timezone.utc),
        updated_at=sample.updated_at or datetime.now(timezone.utc),
    )


def list_samples(db: Session, skip: int = 0, limit: int = 100) -> List[SampleListItem]:
    """List all samples with computed statistics."""
    samples = (
        db.query(Sample)
        .order_by(Sample.updated_at.desc())
        .offset(skip)
        .limit(limit)
        .all()
    )
    return [_sample_to_list_item(s) for s in samples]


def get_sample(db: Session, sample_id: int) -> Optional[Sample]:
    """Get a single sample by ID."""
    return db.query(Sample).filter(Sample.id == sample_id).first()


def get_sample_by_code(db: Session, code: str) -> Optional[Sample]:
    """Get a sample by its unique code."""
    return db.query(Sample).filter(Sample.code == code).first()


def create_sample(db: Session, data: SampleCreate) -> Sample:
    """Create a new sample. Raises 409 if code already exists."""
    existing = get_sample_by_code(db, data.code)
    if existing:
        raise HTTPException(status_code=409, detail=f"样品编号 '{data.code}' 已存在")
    sample = Sample(
        name=data.name,
        code=data.code,
        category=data.category,
        brand=data.brand,
        model=data.model,
        color_name=data.color_name,
        other_info=data.other_info,
        test_condition=data.test_condition,
        aging_time=data.aging_time,
        device_info=data.device_info,
        test_device=data.test_device,
        measurement_test=data.measurement_test,
        description=data.description,
    )
    db.add(sample)
    db.commit()
    db.refresh(sample)
    return sample


def update_sample(db: Session, sample_id: int, data: SampleUpdate) -> Sample:
    """Update sample fields. Raises 404 if not found, 409 if code conflicts."""
    sample = get_sample(db, sample_id)
    if not sample:
        raise HTTPException(status_code=404, detail="样品不存在")

    update_data = data.model_dump(exclude_unset=True)

    # Check code uniqueness if changing code
    if "code" in update_data and update_data["code"] != sample.code:
        existing = get_sample_by_code(db, update_data["code"])
        if existing:
            raise HTTPException(status_code=409, detail=f"样品编号 '{update_data['code']}' 已存在")

    for key, value in update_data.items():
        setattr(sample, key, value)

    db.commit()
    db.refresh(sample)
    return sample


def delete_sample(db: Session, sample_id: int) -> bool:
    """Delete a sample and all related data. Raises 404 if not found."""
    sample = get_sample(db, sample_id)
    if not sample:
        raise HTTPException(status_code=404, detail="样品不存在")

    # Delete photo files from disk
    for photo in sample.photos:
        filepath = os.path.join("uploads", photo.filename)
        if os.path.exists(filepath):
            os.remove(filepath)

    db.delete(sample)
    db.commit()
    return True


# ============================================================
# Measurement CRUD
# ============================================================

def list_measurements(db: Session, sample_id: int) -> List[dict]:
    """List all measurements for a sample, with is_baseline and photo_count."""
    sample = get_sample(db, sample_id)
    if not sample:
        raise HTTPException(status_code=404, detail="样品不存在")

    baselines = get_baseline_measurements(db, sample_id)
    baseline_ids = {b.id for b in baselines}

    measurements = _get_visible_measurements_for_sample(db, sample_id)

    result = []
    for m in measurements:
        result.append({
            "id": m.id,
            "sample_id": m.sample_id,
            "measurement_date": m.measurement_date,
            "device": m.device or "SP64",
            "angle": m.angle,
            "aging_hours": m.aging_hours or 0,
            "L": m.L,
            "a": m.a,
            "b": m.b,
            "delta_E": m.delta_E,
            "notes": m.notes or "",
            "created_at": m.created_at,
            "is_baseline": m.id in baseline_ids,
            "photo_count": len(m.photos),
        })
    return result


def list_all_measurements(db: Session) -> List[dict]:
    """List the latest visible measurements across all samples."""
    samples = (
        db.query(Sample)
        .order_by(Sample.name.asc(), Sample.code.asc(), Sample.id.asc())
        .all()
    )

    result = []
    for sample in samples:
        baselines = get_baseline_measurements(db, sample.id)
        baseline_ids = {b.id for b in baselines}
        measurements = _get_visible_measurements_for_sample(db, sample.id)

        for measurement in measurements:
            result.append({
                "id": measurement.id,
                "sample_id": sample.id,
                "sample_name": sample.name,
                "sample_code": sample.code,
                "sample_brand": sample.brand or "",
                "sample_model": sample.model or "",
                "sample_test_condition": sample.test_condition or "",
                "sample_aging_time": sample.aging_time or "",
                "measurement_date": measurement.measurement_date,
                "device": measurement.device or "SP64",
                "angle": measurement.angle,
                "aging_hours": measurement.aging_hours or 0,
                "L": measurement.L,
                "a": measurement.a,
                "b": measurement.b,
                "delta_E": measurement.delta_E,
                "notes": measurement.notes or "",
                "created_at": measurement.created_at,
                "is_baseline": measurement.id in baseline_ids,
                "photo_count": len(measurement.photos),
            })

    return result


def get_measurement(db: Session, measurement_id: int) -> Optional[ColorMeasurement]:
    """Get a single measurement by ID."""
    return db.query(ColorMeasurement).filter(ColorMeasurement.id == measurement_id).first()


def create_measurement(db: Session, sample_id: int, data: MeasurementCreate) -> ColorMeasurement:
    """Create a measurement with automatic delta_E calculation.
    aging_hours=0 → baseline (delta_E=null)."""
    sample = get_sample(db, sample_id)
    if not sample:
        raise HTTPException(status_code=404, detail="样品不存在")

    measurement_date = data.measurement_date or beijing_now()
    existing_measurements = (
        db.query(ColorMeasurement)
        .filter(ColorMeasurement.sample_id == sample_id)
        .all()
    )
    if _has_exact_duplicate_measurement(
        existing_measurements,
        sample_id,
        data.device,
        data.angle,
        data.aging_hours,
        data.L,
        data.a,
        data.b,
        data.notes,
    ):
        raise HTTPException(status_code=409, detail="存在完全相同的测量数据，未重复保存")

    measurement = ColorMeasurement(
        sample_id=sample_id,
        measurement_date=measurement_date,
        device=data.device,
        angle=data.angle,
        aging_hours=data.aging_hours,
        L=data.L,
        a=data.a,
        b=data.b,
        delta_E=None,
        notes=data.notes,
    )
    db.add(measurement)
    db.commit()
    db.refresh(measurement)

    # Recalculate delta_E for all (new baseline might change others)
    recalculate_delta_e_for_sample(db, sample_id)
    db.refresh(measurement)
    return measurement


def create_measurements_batch(db: Session, sample_id: int,
                               date_val: datetime, device: str,
                               aging_hours: float, notes: str,
                               entries: list) -> List[ColorMeasurement]:
    """Batch create measurements (for MT12 multi-angle).
    aging_hours=0 → all 6 angles are baselines."""
    sample = get_sample(db, sample_id)
    if not sample:
        raise HTTPException(status_code=404, detail="样品不存在")

    existing_measurements = (
        db.query(ColorMeasurement)
        .filter(ColorMeasurement.sample_id == sample_id)
        .all()
    )
    duplicate_angles = []
    for entry in entries:
        if _has_exact_duplicate_measurement(
            existing_measurements,
            sample_id,
            device,
            entry.angle,
            aging_hours,
            entry.L,
            entry.a,
            entry.b,
            notes,
        ):
            duplicate_angles.append(entry.angle or "SP64")

    if duplicate_angles:
        raise HTTPException(
            status_code=409,
            detail=f"存在完全相同的测量数据，未重复保存: {', '.join(duplicate_angles)}",
        )

    results = []
    for entry in entries:
        measurement = ColorMeasurement(
            sample_id=sample_id,
            measurement_date=date_val,
            device=device,
            angle=entry.angle,
            aging_hours=aging_hours,
            L=entry.L,
            a=entry.a,
            b=entry.b,
            delta_E=None,  # Will be calculated below
            notes=notes,
        )
        db.add(measurement)
        results.append(measurement)

    db.commit()
    for m in results:
        db.refresh(m)

    # Recalculate delta_E (aging_hours=0 → baseline, others → compare)
    recalculate_delta_e_for_sample(db, sample_id)
    for m in results:
        db.refresh(m)
    return results


def update_measurement(db: Session, measurement_id: int, data: MeasurementUpdate) -> ColorMeasurement:
    """Update a measurement. Recalculates delta_E for the sample after update."""
    measurement = get_measurement(db, measurement_id)
    if not measurement:
        raise HTTPException(status_code=404, detail="测量记录不存在")

    update_data = data.model_dump(exclude_unset=True)
    old_aging = measurement.aging_hours

    for key, value in update_data.items():
        setattr(measurement, key, value)

    db.commit()

    # Always recalculate delta_E for the sample (aging_hours may have changed)
    recalculate_delta_e_for_sample(db, measurement.sample_id)

    db.refresh(measurement)
    return measurement


def delete_measurement(db: Session, measurement_id: int) -> bool:
    """Delete a measurement. If it was a baseline, recalculate all remaining."""
    measurement = get_measurement(db, measurement_id)
    if not measurement:
        raise HTTPException(status_code=404, detail="测量记录不存在")

    sample_id = measurement.sample_id
    was_baseline = (measurement.aging_hours == 0)

    # Delete associated photo files
    for photo in measurement.photos:
        filepath = os.path.join("uploads", photo.filename)
        if os.path.exists(filepath):
            os.remove(filepath)

    db.delete(measurement)
    db.commit()

    if was_baseline:
        recalculate_delta_e_for_sample(db, sample_id)

    return True


# ============================================================
# Photo CRUD
# ============================================================

ALLOWED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB


def create_photo(db: Session, sample_id: int, file: UploadFile,
                 measurement_id: Optional[int] = None, notes: str = "") -> Photo:
    """Upload and create a photo record."""
    sample = get_sample(db, sample_id)
    if not sample:
        raise HTTPException(status_code=404, detail="样品不存在")

    # Validate file extension
    _, ext = os.path.splitext(file.filename or "")
    ext_lower = ext.lower()
    if ext_lower not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"不支持的文件格式 '{ext}'。支持的格式: {', '.join(ALLOWED_EXTENSIONS)}"
        )

    # Validate measurement if provided
    if measurement_id is not None:
        measurement = get_measurement(db, measurement_id)
        if not measurement or measurement.sample_id != sample_id:
            raise HTTPException(status_code=404, detail="测量记录不存在或不属于该样品")

    # Generate unique filename
    unique_name = f"{uuid.uuid4().hex}{ext_lower}"
    os.makedirs("uploads", exist_ok=True)
    filepath = os.path.join("uploads", unique_name)

    # Read and save file
    contents = file.file.read()
    if len(contents) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="文件大小超过 10MB 限制")

    with open(filepath, "wb") as f:
        f.write(contents)

    photo = Photo(
        sample_id=sample_id,
        measurement_id=measurement_id,
        filename=unique_name,
        original_name=file.filename or "unknown",
        notes=notes,
    )
    db.add(photo)
    db.commit()
    db.refresh(photo)
    return photo


def get_photo(db: Session, photo_id: int) -> Optional[Photo]:
    """Get photo metadata by ID."""
    return db.query(Photo).filter(Photo.id == photo_id).first()


def list_photos_by_sample(db: Session, sample_id: int) -> List[Photo]:
    """List all photos for a sample."""
    return (
        db.query(Photo)
        .filter(Photo.sample_id == sample_id)
        .order_by(Photo.upload_date.desc())
        .all()
    )


def list_photos_by_measurement(db: Session, measurement_id: int) -> List[Photo]:
    """List photos linked to a specific measurement."""
    return (
        db.query(Photo)
        .filter(Photo.measurement_id == measurement_id)
        .order_by(Photo.upload_date.desc())
        .all()
    )


def delete_photo(db: Session, photo_id: int) -> bool:
    """Delete a photo record and its file."""
    photo = get_photo(db, photo_id)
    if not photo:
        raise HTTPException(status_code=404, detail="照片不存在")

    filepath = os.path.join("uploads", photo.filename)
    if os.path.exists(filepath):
        os.remove(filepath)

    db.delete(photo)
    db.commit()
    return True


# ============================================================
# Search
# ============================================================

def search_samples(db: Session, q: Optional[str] = None,
                   start_date: Optional[date] = None,
                   end_date: Optional[date] = None,
                   skip: int = 0, limit: int = 100) -> Tuple[List[SampleListItem], int]:
    """Search samples by keyword, date range. Returns (results, total_count)."""
    query = db.query(Sample)

    if q:
        search_term = f"%{q}%"
        query = query.filter(
            or_(
                Sample.name.ilike(search_term),
                Sample.code.ilike(search_term),
                Sample.category.ilike(search_term),
                Sample.brand.ilike(search_term),
                Sample.model.ilike(search_term),
                Sample.color_name.ilike(search_term),
                Sample.test_condition.ilike(search_term),
                Sample.aging_time.ilike(search_term),
                Sample.description.ilike(search_term),
            )
        )

    if start_date or end_date:
        # Filter samples that have measurements within the date range
        meas_filters = []
        if start_date:
            meas_filters.append(ColorMeasurement.measurement_date >= start_date)
        if end_date:
            meas_filters.append(ColorMeasurement.measurement_date <= end_date)
        if meas_filters:
            subquery = (
                db.query(ColorMeasurement.sample_id)
                .filter(and_(*meas_filters))
                .distinct()
                .subquery()
            )
            query = query.filter(Sample.id.in_(subquery))

    total = query.count()
    samples = query.order_by(Sample.updated_at.desc()).offset(skip).limit(limit).all()

    return [_sample_to_list_item(s) for s in samples], total


# ============================================================
# Chart Data
# ============================================================

def get_chart_data(db: Session, sample_id: int) -> dict:
    """Get chart-ready data: 4 separate charts (L*, a*, b*, ΔE) over aging_hours.
    Groups by device+angle, returns one series per device/angle combination."""
    sample = get_sample(db, sample_id)
    if not sample:
        raise HTTPException(status_code=404, detail="样品不存在")

    measurements = _get_visible_measurements_for_sample(db, sample_id)

    # Group by device+angle to create separate lines with one point per aging hour.
    series_map = {}
    for m in measurements:
        key = f"{_normalize_device(m.device)}|{_normalize_angle(m.angle)}"
        if key not in series_map:
            series_map[key] = {"points": {}}
        aging_hours = _normalize_aging_hours(m.aging_hours)
        series_map[key]["points"][aging_hours] = {
            "L": m.L,
            "a": m.a,
            "b": m.b,
            "de": m.delta_E if m.delta_E is not None else 0,
        }

    common_labels = sorted({_normalize_aging_hours(m.aging_hours) for m in measurements})

    angle_colors = {
        "": "rgb(37, 99, 235)",
        "r45as-15": "rgb(239, 68, 68)",
        "r45as15": "rgb(245, 158, 11)",
        "r45as25": "rgb(34, 197, 94)",
        "r45as45": "rgb(59, 130, 246)",
        "r45as75": "rgb(168, 85, 247)",
        "r45as110": "rgb(236, 72, 153)",
    }

    charts = {
        "L": {"title": "L* 趋势", "yLabel": "L*", "datasets": []},
        "a": {"title": "a* 趋势", "yLabel": "a*", "datasets": []},
        "b": {"title": "b* 趋势", "yLabel": "b*", "datasets": []},
        "delta_E": {"title": "ΔE 趋势", "yLabel": "ΔE", "datasets": []},
    }

    for key, series in series_map.items():
        device, angle = key.split("|")
        label = f"{device}" if not angle else f"{device} {angle}"
        c = angle_colors.get(angle, "rgb(107, 114, 128)")
        aligned_L = [series["points"].get(h, {}).get("L") for h in common_labels]
        aligned_a = [series["points"].get(h, {}).get("a") for h in common_labels]
        aligned_b = [series["points"].get(h, {}).get("b") for h in common_labels]
        aligned_de = [series["points"].get(h, {}).get("de") for h in common_labels]

        charts["L"]["datasets"].append({
            "label": label, "data": aligned_L,
            "borderColor": c, "backgroundColor": c, "tension": 0.3,
            "pointRadius": 3,
            "spanGaps": False,
        })
        charts["a"]["datasets"].append({
            "label": label, "data": aligned_a,
            "borderColor": c, "backgroundColor": c, "tension": 0.3,
            "pointRadius": 3,
            "spanGaps": False,
        })
        charts["b"]["datasets"].append({
            "label": label, "data": aligned_b,
            "borderColor": c, "backgroundColor": c, "tension": 0.3,
            "pointRadius": 3,
            "spanGaps": False,
        })
        charts["delta_E"]["datasets"].append({
            "label": label, "data": aligned_de,
            "borderColor": c, "backgroundColor": c, "tension": 0.3,
            "pointRadius": 3,
            "spanGaps": False,
        })

    return {
        "labels": common_labels,
        "charts": charts,
    }


# ============================================================
# Import (shared processing)
# ============================================================

def _parse_import_rows(header: list, data_rows: list) -> tuple[dict, list]:
    """Parse import rows into grouped sample data and collect row-level errors."""
    col_map = {}
    expected_cols = {
        "样品名称": "name",
        "样品编号": "code",
        "类别": "category",
        "品牌": "brand",
        "型号": "model",
        "颜色": "color_name",
        "测试条件": "test_condition",
        "老化时间": "aging_time",
        "样品描述": "description",
        "测量日期": "date",
        "测试设备": "device",
        "角度": "angle",
        "测量老化时间(小时)": "aging_hours",
        "L*": "L",
        "a*": "a",
        "b*": "b",
        "备注": "notes",
    }

    for idx, h in enumerate(header):
        h_clean = str(h).strip() if h else ""
        if h_clean in expected_cols:
            col_map[expected_cols[h_clean]] = idx

    # Required columns
    required = ["name", "code", "L", "a", "b"]
    missing = [r for r in required if r not in col_map]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"文件缺少必要列: {', '.join(missing)}。"
                   f"需要: 样品名称, 样品编号, L*, a*, b*"
        )

    errors = []

    # Group rows by sample code
    sample_rows = {}
    for row_num, row in enumerate(data_rows, start=2):
        code = str(row[col_map["code"]]).strip() if row[col_map["code"]] else ""
        if not code:
            errors.append(f"第{row_num}行: 样品编号为空，跳过")
            continue

        name = str(row[col_map["name"]]).strip() if row[col_map["name"]] else ""
        if not name:
            errors.append(f"第{row_num}行: 样品名称为空，跳过")
            continue

        # Validate L, a, b
        try:
            L_val = float(row[col_map["L"]]) if row[col_map["L"]] is not None else None
            a_val = float(row[col_map["a"]]) if row[col_map["a"]] is not None else None
            b_val = float(row[col_map["b"]]) if row[col_map["b"]] is not None else None
        except (ValueError, TypeError):
            errors.append(f"第{row_num}行: L*/a*/b* 值格式错误，跳过")
            continue

        if L_val is None or a_val is None or b_val is None:
            errors.append(f"第{row_num}行: L*/a*/b* 值不完整，跳过")
            continue

        if L_val < 0 or L_val > 200:
            errors.append(f"第{row_num}行: L* 值 {L_val} 超出 0-200 范围，跳过")
            continue

        # Parse date
        date_val = None
        if "date" in col_map and len(row) > col_map["date"] and row[col_map["date"]] is not None:
            raw_date = row[col_map["date"]]
            if isinstance(raw_date, str):
                raw_date = raw_date.strip()
                for fmt in ["%Y-%m-%d %H:%M", "%Y-%m-%d", "%Y/%m/%d %H:%M", "%Y/%m/%d",
                           "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"]:
                    try:
                        date_val = datetime.strptime(raw_date, fmt)
                        break
                    except ValueError:
                        continue
                if date_val is None and raw_date:
                    errors.append(f"第{row_num}行: 日期格式无法识别 '{raw_date}'，使用当前时间")
            elif isinstance(raw_date, datetime):
                date_val = raw_date
            else:
                date_val = beijing_now()
        else:
            date_val = beijing_now()

        if date_val is None:
            date_val = beijing_now()

        description = ""
        if "description" in col_map and len(row) > col_map["description"] and row[col_map["description"]] is not None:
            description = str(row[col_map["description"]]).strip()

        category = ""
        if "category" in col_map and len(row) > col_map["category"] and row[col_map["category"]] is not None:
            category = str(row[col_map["category"]]).strip()

        brand = ""
        if "brand" in col_map and len(row) > col_map["brand"] and row[col_map["brand"]] is not None:
            brand = str(row[col_map["brand"]]).strip()

        model = ""
        if "model" in col_map and len(row) > col_map["model"] and row[col_map["model"]] is not None:
            model = str(row[col_map["model"]]).strip()

        color_name = ""
        if "color_name" in col_map and len(row) > col_map["color_name"] and row[col_map["color_name"]] is not None:
            color_name = str(row[col_map["color_name"]]).strip()

        test_condition = ""
        if "test_condition" in col_map and len(row) > col_map["test_condition"] and row[col_map["test_condition"]] is not None:
            test_condition = str(row[col_map["test_condition"]]).strip()

        aging_time = ""
        if "aging_time" in col_map and len(row) > col_map["aging_time"] and row[col_map["aging_time"]] is not None:
            aging_time = str(row[col_map["aging_time"]]).strip()

        aging_hours = 0.0
        aging_hours_raw = None
        if "aging_hours" in col_map and len(row) > col_map["aging_hours"]:
            aging_hours_raw = row[col_map["aging_hours"]]

        aging_hours_missing = aging_hours_raw in (None, "")
        if not aging_hours_missing:
            try:
                aging_hours = float(aging_hours_raw)
            except (ValueError, TypeError):
                aging_hours = 0.0
                aging_hours_missing = True

        # Backward-compatible fallback for CSVs that still place measurement aging
        # hours into the sample aging_time column after the template column order changed.
        if aging_hours_missing and aging_time:
            try:
                aging_hours = float(aging_time)
                errors.append(f"第{row_num}行: 未填写'测量老化时间(小时)'，已使用'老化时间'列值 {aging_time}")
            except (ValueError, TypeError):
                aging_hours = 0.0

        notes = ""
        if "notes" in col_map and len(row) > col_map["notes"] and row[col_map["notes"]] is not None:
            notes = str(row[col_map["notes"]]).strip()

        device = "SP64"
        if "device" in col_map and len(row) > col_map["device"] and row[col_map["device"]] is not None:
            device = str(row[col_map["device"]]).strip() or "SP64"

        angle = None
        if "angle" in col_map and len(row) > col_map["angle"] and row[col_map["angle"]] is not None:
            angle_val = str(row[col_map["angle"]]).strip()
            if angle_val:
                angle = angle_val

        # Unique key = code + name
        sample_key = f"{code}|{name}"
        if sample_key not in sample_rows:
            sample_rows[sample_key] = {
                "code": code, "name": name, "description": description,
                "category": category, "brand": brand, "model": model,
                "color_name": color_name, "test_condition": test_condition,
                "aging_time": aging_time, "test_device": device, "measurements": []
            }
        elif not sample_rows[sample_key].get("test_device") and device:
            sample_rows[sample_key]["test_device"] = device
        sample_rows[sample_key]["measurements"].append({
            "row_num": row_num,
            "date": date_val,
            "device": device,
            "angle": angle,
            "aging_hours": aging_hours,
            "L": L_val,
            "a": a_val,
            "b": b_val,
            "notes": notes,
        })

    return sample_rows, errors


def _build_import_preview_from_rows(header: list, data_rows: list) -> dict:
    sample_rows, errors = _parse_import_rows(header, data_rows)
    samples = [
        {
            "name": data["name"],
            "code": data["code"],
            "measurement_count": len(data["measurements"]),
        }
        for data in sample_rows.values()
    ]
    return {
        "total_samples": len(sample_rows),
        "total_measurements": sum(sample["measurement_count"] for sample in samples),
        "samples": samples,
        "errors": errors,
    }


def _process_import_rows(db: Session, header: list, data_rows: list) -> dict:
    """Shared import logic: parse header, validate rows, create samples & measurements."""
    sample_rows, errors = _parse_import_rows(header, data_rows)
    samples_created = 0
    measurements_created = 0

    # Create samples and measurements
    for sample_key, data in sample_rows.items():
        code = data["code"]
        name = data["name"]

        # Match by code AND name (unique identifier)
        sample = db.query(Sample).filter(
            Sample.code == code, Sample.name == name
        ).first()

        if not sample:
            sample = Sample(
                name=name,
                code=code,
                category=data.get("category", ""),
                brand=data.get("brand", ""),
                model=data.get("model", ""),
                color_name=data.get("color_name", ""),
                test_condition=data.get("test_condition", ""),
                aging_time=data.get("aging_time", ""),
                test_device=data.get("test_device", ""),
                description=data["description"],
            )
            db.add(sample)
            db.flush()
            samples_created += 1
        else:
            if not sample.description and data["description"]:
                sample.description = data["description"]
            for field in ["category", "brand", "model", "color_name", "test_condition", "aging_time", "test_device"]:
                existing_val = getattr(sample, field, "") or ""
                new_val = data.get(field, "")
                if not existing_val and new_val:
                    setattr(sample, field, new_val)

        existing_measurements = (
            db.query(ColorMeasurement)
            .filter(ColorMeasurement.sample_id == sample.id)
            .all()
        )
        existing_by_key = defaultdict(list)
        for existing in existing_measurements:
            existing_by_key[_measurement_key(sample.id, existing.device, existing.angle, existing.aging_hours)].append(existing)

        # Keep import order for same-key collisions so the last imported row becomes the visible one.
        data["measurements"].sort(key=lambda m: (m.get("aging_hours", 0), m["date"], m["row_num"]))

        for m_data in data["measurements"]:
            measurement_key = _measurement_key(
                sample.id,
                m_data.get("device", "SP64"),
                m_data.get("angle"),
                m_data.get("aging_hours", 0),
            )
            measurement_signature = _measurement_signature(
                m_data["L"],
                m_data["a"],
                m_data["b"],
                m_data["notes"],
            )

            is_exact_duplicate = any(
                _measurement_signature(existing.L, existing.a, existing.b, existing.notes) == measurement_signature
                for existing in existing_by_key[measurement_key]
            )
            if is_exact_duplicate:
                angle_label = m_data.get("angle") or "-"
                errors.append(
                    f"第{m_data['row_num']}行: {name} / {code} 在 {m_data.get('aging_hours', 0)}h / {m_data.get('device', 'SP64')} / {angle_label} 的数据重复，已跳过"
                )
                continue

            measurement = ColorMeasurement(
                sample_id=sample.id,
                measurement_date=m_data["date"],
                device=m_data.get("device", "SP64"),
                angle=m_data.get("angle"),
                aging_hours=m_data.get("aging_hours", 0),
                L=m_data["L"],
                a=m_data["a"],
                b=m_data["b"],
                notes=m_data["notes"],
            )
            db.add(measurement)
            existing_by_key[measurement_key].append(measurement)
            measurements_created += 1

    db.commit()

    for sample_key in sample_rows:
        d = sample_rows[sample_key]
        sample = db.query(Sample).filter(
            Sample.code == d["code"], Sample.name == d["name"]
        ).first()
        if sample:
            recalculate_delta_e_for_sample(db, sample.id)

    return {
        "samples_created": samples_created,
        "samples_updated": len(sample_rows) - samples_created,
        "measurements_created": measurements_created,
        "total_samples": len(sample_rows),
        "errors": errors,
    }


def import_from_excel(db: Session, file_path: str) -> dict:
    """Import from .xlsx file. Reads row 1 as header, processes data rows."""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel 文件为空或只有表头")

    header = [str(h).strip() if h else "" for h in rows[0]]
    data_rows = rows[1:]
    return _process_import_rows(db, header, data_rows)


def import_from_csv(db: Session, file_path: str) -> dict:
    """Import from .csv file. Reads row 1 as header, processes data rows.
    Supports UTF-8, GBK, GB2312, GB18030 encodings."""
    import csv

    # Try multiple encodings (common for Chinese CSV files)
    encodings = ["utf-8-sig", "utf-8", "gbk", "gb18030", "gb2312", "latin-1"]
    all_rows = None
    last_error = None

    for enc in encodings:
        try:
            with open(file_path, "r", encoding=enc) as f:
                reader = csv.reader(f)
                all_rows = list(reader)
            break  # Success
        except (UnicodeDecodeError, UnicodeError) as e:
            last_error = str(e)
            continue

    if all_rows is None:
        raise HTTPException(
            status_code=400,
            detail=f"无法识别 CSV 文件编码，请将文件另存为 UTF-8 格式。错误: {last_error}"
        )

    if len(all_rows) < 2:
        raise HTTPException(status_code=400, detail="CSV 文件为空或只有表头")

    header = all_rows[0]
    data_rows = all_rows[1:]
    return _process_import_rows(db, header, data_rows)


def preview_import_from_excel(file_path: str) -> dict:
    """Preview import summary from .xlsx file without writing to the database."""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel 文件为空或只有表头")

    header = [str(h).strip() if h else "" for h in rows[0]]
    data_rows = rows[1:]
    return _build_import_preview_from_rows(header, data_rows)


def preview_import_from_csv(file_path: str) -> dict:
    """Preview import summary from .csv file without writing to the database."""
    import csv

    encodings = ["utf-8-sig", "utf-8", "gbk", "gb18030", "gb2312", "latin-1"]
    all_rows = None
    last_error = None

    for enc in encodings:
        try:
            with open(file_path, "r", encoding=enc) as f:
                reader = csv.reader(f)
                all_rows = list(reader)
            break
        except (UnicodeDecodeError, UnicodeError) as e:
            last_error = str(e)
            continue

    if all_rows is None:
        raise HTTPException(
            status_code=400,
            detail=f"无法识别 CSV 文件编码，请将文件另存为 UTF-8 格式。错误: {last_error}"
        )

    if len(all_rows) < 2:
        raise HTTPException(status_code=400, detail="CSV 文件为空或只有表头")

    header = all_rows[0]
    data_rows = all_rows[1:]
    return _build_import_preview_from_rows(header, data_rows)


def preview_import_file(file_path: str, filename: str) -> dict:
    """Auto-detect file type and build an import preview without persisting data."""
    _, ext = os.path.splitext(filename)
    ext = ext.lower()
    if ext in {".xlsx", ".xls"}:
        return preview_import_from_excel(file_path)
    if ext == ".csv":
        return preview_import_from_csv(file_path)
    raise HTTPException(status_code=400, detail=f"不支持的文件格式 '{ext}'。支持: .xlsx, .xls, .csv")


def import_file(db: Session, file_path: str, filename: str) -> dict:
    """Auto-detect file type and import. Supports .xlsx, .xls, .csv."""
    _, ext = os.path.splitext(filename)
    ext = ext.lower()
    if ext in {".xlsx", ".xls"}:
        return import_from_excel(db, file_path)
    elif ext == ".csv":
        return import_from_csv(db, file_path)
    else:
        raise HTTPException(status_code=400, detail=f"不支持的文件格式 '{ext}'。支持: .xlsx, .xls, .csv")


# ============================================================
# Export
# ============================================================

def export_to_excel(db: Session, sample_ids: Optional[List[int]] = None) -> bytes:
    """Export measurement data to Excel format. Returns bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "颜色老化数据"

    # Header style
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "样品名称", "样品编号", "类别", "品牌", "型号", "颜色",
        "测试条件", "老化时间", "样品描述",
        "测量日期", "测试设备", "角度", "测量老化时间(小时)", "L*", "a*", "b*", "ΔE", "备注"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data
    query = db.query(Sample)
    if sample_ids:
        query = query.filter(Sample.id.in_(sample_ids))

    row = 2
    for sample in query.order_by(Sample.code).all():
        measurements = (
            db.query(ColorMeasurement)
            .filter(ColorMeasurement.sample_id == sample.id)
            .order_by(ColorMeasurement.measurement_date)
            .all()
        )
        for m in measurements:
            ws.cell(row=row, column=1, value=sample.name).border = thin_border
            ws.cell(row=row, column=2, value=sample.code).border = thin_border
            ws.cell(row=row, column=3, value=sample.category or "").border = thin_border
            ws.cell(row=row, column=4, value=sample.brand or "").border = thin_border
            ws.cell(row=row, column=5, value=sample.model or "").border = thin_border
            ws.cell(row=row, column=6, value=sample.color_name or "").border = thin_border
            ws.cell(row=row, column=7, value=sample.test_condition or "").border = thin_border
            ws.cell(row=row, column=8, value=sample.aging_time or "").border = thin_border
            ws.cell(row=row, column=9, value=sample.description or "").border = thin_border
            ws.cell(row=row, column=10, value=m.measurement_date.strftime("%Y-%m-%d %H:%M")).border = thin_border
            ws.cell(row=row, column=11, value=m.device or "SP64").border = thin_border
            ws.cell(row=row, column=12, value=m.angle or "").border = thin_border
            ws.cell(row=row, column=13, value=m.aging_hours or 0).border = thin_border
            ws.cell(row=row, column=14, value=m.L).border = thin_border
            ws.cell(row=row, column=15, value=m.a).border = thin_border
            ws.cell(row=row, column=16, value=m.b).border = thin_border
            ws.cell(row=row, column=17, value=m.delta_E).border = thin_border
            ws.cell(row=row, column=18, value=m.notes or "").border = thin_border
            row += 1
        if not measurements:
            # Sample with no measurements
            ws.cell(row=row, column=1, value=sample.name).border = thin_border
            ws.cell(row=row, column=2, value=sample.code).border = thin_border
            ws.cell(row=row, column=3, value=sample.category or "").border = thin_border
            ws.cell(row=row, column=4, value=sample.brand or "").border = thin_border
            ws.cell(row=row, column=5, value=sample.model or "").border = thin_border
            ws.cell(row=row, column=6, value=sample.color_name or "").border = thin_border
            ws.cell(row=row, column=7, value=sample.test_condition or "").border = thin_border
            ws.cell(row=row, column=8, value=sample.aging_time or "").border = thin_border
            ws.cell(row=row, column=9, value=sample.description or "").border = thin_border
            for c in range(10, 19):
                ws.cell(row=row, column=c, value="").border = thin_border
            row += 1

    # Column widths
    widths = [20, 15, 12, 15, 15, 15, 20, 15, 30, 18, 10, 12, 14, 8, 8, 8, 8, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
